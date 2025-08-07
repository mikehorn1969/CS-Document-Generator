from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from flask import send_file


def create_tables():

    # Example: two DataFrames
    df1 = pd.DataFrame({
        'Name': ['Alice', 'Bob'],
        'Score': [90, 85]
    })

    df2 = pd.DataFrame({
        'Department': ['IT', 'HR'],
        'Employees': [12, 7]
    })

    # Write both DataFrames to different row positions in the same sheet
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write df1 at row 0
        df1.to_excel(writer, index=False, startrow=0, sheet_name='Sheet1')

        # Write df2 below df1 + some space
        df2_start_row = len(df1) + 3  # 3 = 1 for header + 2 spacing
        df2.to_excel(writer, index=False, startrow=df2_start_row, sheet_name='Sheet1')        

    # Reopen to add tables
    output.seek(0)
    wb = load_workbook(output)
    ws = wb['Sheet1']

    # ===== Add Table 1 =====
    df1_rows = len(df1) + 1
    df1_cols = len(df1.columns)
    df1_range = f"A1:{get_column_letter(df1_cols)}{df1_rows}"

    table1 = Table(displayName="Table1", ref=df1_range)
    table1.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table1)

    # ===== Add Table 2 =====
    df2_rows = len(df2) + 1
    df2_cols = len(df2.columns)
    start_row2 = len(df1) + 4  # Where we started writing df2
    start_cell2 = f"A{start_row2}"
    end_cell2 = f"{get_column_letter(df2_cols)}{start_row2 + len(df2)}"
    df2_range = f"{start_cell2}:{end_cell2}"

    table2 = Table(displayName="Table2", ref=df2_range)
    table2.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table2)

    # Save final output
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    # Return download
    return send_file(
        final_output,
        as_attachment=True,
        download_name="multi_table_export.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


if __name__ == "__main__":
    result = create_tables()