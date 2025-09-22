import os

from app import db

from flask import render_template, request, redirect, url_for, session, send_file, flash, jsonify
from flask import Blueprint

from app.models import ServiceStandard, ServiceArrangement, ServiceContract as ContractModel
from app.c7query import  searchC7Candidate, loadC7Clients, getContactsByCompany, gather_data, getC7Candidate, loadServiceStandards, loadServiceArrangements, getC7Candidates
from app.chquery import validateCH, searchCH
from app.classes import Company, Config
from app.helper import formatName, uploadToSharePoint
from datetime import datetime
from sqlalchemy import select
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import time
from typing import List

views_bp = Blueprint('views', __name__)


@views_bp.route('/', methods=["GET", "POST"])
def index():
    return render_template('index.html', sid=session.get('sid', ''))


@views_bp.route('/searchcandidates')
def search_candidates():
    q = request.args.get("q", "").strip()
    results = fetch_candidates(q)
    session["candidateName"] = results[0].get("candidateName") if results else None
    return jsonify(results)
    

@views_bp.route('/searchclients')
def search_clients():
    q = request.args.get("q", "").strip()
    results = fetch_clients(q)
    return jsonify(results)


@views_bp.route('/searchcontacts')
def search_contacts():
    qclient = request.args.get('client','').strip()
    qcontact = request.args.get("q", "").strip()
    results = fetch_contacts(qclient,qcontact)
    return jsonify(results)

          
@views_bp.route('/clearsession', methods=["POST"])
def clear_session():
    session.clear()
    return redirect(url_for('views.index'))


@views_bp.route('/colleaguedata', methods=["GET"])
def colleaguedata():
    """
    Handles GET requests to load and prepare contract data for the colleague view.
    This function attempts to retrieve contract data in the following order:
    1. From the session (if previously saved).
    2. From the database using the session's 'sid'.
    3. From an external C7 source if not found in session or database.
    If data is loaded from C7, it maps and formats relevant fields, including calculating the contract duration based on start and end dates. The contract data is then stored in the session and passed to the 'colleague.html' template for rendering.
    Returns:
        Rendered HTML template 'colleague.html' with contract data context.
    """
    session_contract = session.get('sessionContract') or None
    if not session_contract:
        flash("Select a Service Provider before continuing.", "error")
        return redirect(url_for('views.index'))
    
    return render_template(
        'colleague.html', contractdata=session_contract)


@views_bp.route('/validate_c7', methods=["POST"])
def validate_c7():
    """
    Validates client and service provider details against Companies House
    """

    contract = session.get('sessionContract', {})
    passed = True
    
    if contract:
        # validate data against CH records
        # 1. Candidate        
        ch_candidatename = contract.get("candidateName").split("(")
        ch_candidatename = ch_candidatename[0]
        ch_result = validateCH(contract.get("candidateltdregno"), contract.get("candidateltdname"), ch_candidatename)

        if not ch_result.get("Valid", False):
            flash(ch_result.get("Narrative",""), "error")
            passed = False

        # 2. Client - only where clientname is present
        client_companyname = contract.get("companyregistrationnumber")

        if client_companyname:
            ch_result = validateCH(client_companyname, contract.get("companyname"))

            if not ch_result.get("Valid", False):
                flash(ch_result.get("Narrative",""), "error")
                passed = False

    if passed:
       flash("Companies House validation passed.", "success")

    return redirect(url_for('views.colleaguedata'))
        

@views_bp.route('/servicestandards', methods=['GET', 'POST'])
def set_servicestandards():
    
    standards = []
    session_contract = {}
    service_id = ""
    contract = {}

    # Determine which standards to show - either passed in query string or form data
    which = request.args.get('which')  # "CS Standards" or "SP Standards"
    if not which:
        which = request.form.get('which', 'CS Standards')

    # If SP Standards button clicked, get session data
    if which == "SP Standards":        
        session_contract = session.get('sessionContract', {})
        service_id = session_contract.get("sid", "")
        if service_id is None or service_id == "":
            flash("Select a Service Provider with a Service ID before continuing.", "error")
            return redirect(url_for('views.index'))
    # otherwise we only need the service ID
    else:
        service_id = "CS"

    if request.method == "GET":
                    
        if which == "SP Standards":
            # Load available contract data
            contract = session_contract
            if contract:
                session['sessionContract'] = contract                
        else:
            # No contract data needed for CS Standards
            contract = {"sid": service_id}

    elif request.method == 'POST':

        # Gather standards data
        stdids = request.form.getlist('id')
        ssns = request.form.getlist('ssn')
        descriptions = request.form.getlist('service-description')

        # Zip and process them here:
        standards = list(zip(stdids, ssns, descriptions))

        # Example: print or save them
        for stdid, ssn, desc in standards:
            if ssn:
                rawDesc = desc.strip().strip('"')    
                record = None
                if stdid and stdid.isdigit():
                    record = ServiceStandard.query.get(int(stdid))  
                if record:
                    record.ssn = ssn.strip()
                    record.description = rawDesc
                else:   
                    if ssn.strip() and desc.strip():
                        new_standard = ServiceStandard()
                        new_standard.sid = service_id
                        new_standard.ssn = ssn.strip()
                        new_standard.description = rawDesc
                        db.session.add(new_standard)
        
        db.session.commit()

        # Store standards in session for later use
        #session['serviceStandards'] = [s.to_dict() for s in standards]
        session['serviceStandards'] = [
            {'id': stdid, 'ssn': ssn, 'description': desc.strip().strip('"')}
            for stdid, ssn, desc in standards
        ]

    # (re)load for display 
    standards = loadServiceStandards(service_id) 
    return render_template('standards.html', serviceid=service_id, standards=standards, which=which)


@views_bp.route('/delete/<int:stdid>', methods=['POST'])
def delete_standard(stdid):
    standard = ServiceStandard.query.get_or_404(stdid)
    db.session.delete(standard)
    db.session.commit()
    return redirect(url_for('views.set_servicestandards'))


@views_bp.route('/servicearrangements', methods=['GET', 'POST'])
def manage_servicearrangements():
    
    session_contract = session.get('sessionContract', {})
    service_id = session_contract.get("sid", "")
    if service_id is None or service_id == "":
        flash("Select a Service Provider with a Service ID before continuing.", "error")
        return redirect(url_for('views.index'))

    # Load contract & current arrangements
    contract = gather_data(session_contract)
    if not contract:
        flash("Failed to load contract data.", "error")
        return redirect(url_for('views.index'))

    session['sessionContract'] = contract
    #session["candidateName"] = contract.get("candidateName", "")
    service_id = contract.get("sid", "")

    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

    if request.method == 'POST':
        # For each day, get or create, then update from form (or sensible defaults for new rows)
        for day in days:
            arrangement = (db.session.query(ServiceArrangement)
                           .filter(ServiceArrangement.sid == service_id,
                                   ServiceArrangement.day == day)
                           .one_or_none())

            is_weekend = day in {'Saturday', 'Sunday'}

            def _default(field: str) -> str:
                if is_weekend:
                    if field == 'defaultserviceperiod':
                        return 'As agreed if required'
                    if field == 'atservicebase':
                        return 'As agreed if required'
                    if field == 'atclientlocation':
                        return 'As agreed if required'
                    if field == 'atotherlocation':
                        return 'Prior approval required'
                else:
                    if field == 'defaultserviceperiod':
                        return 'As specified 0800 - 1800'
                    if field == 'atservicebase':
                        return 'As specified'
                    if field == 'atclientlocation':
                        return 'As specified'
                    if field == 'atotherlocation':
                        return 'Prior approval required'
                return ''

            if arrangement is None:
                arrangement = ServiceArrangement(sid=service_id, day=day)
                db.session.add(arrangement)

            # Update from form (fall back to defaults when creating; fall back to existing otherwise)
            arrangement.defaultserviceperiod = request.form.get(f'{day}_default',
                                                                arrangement.defaultserviceperiod or _default('defaultserviceperiod'))
            arrangement.atservicebase = request.form.get(f'{day}_base',
                                                         arrangement.atservicebase or _default('atservicebase'))
            arrangement.atclientlocation = request.form.get(f'{day}_client',
                                                            arrangement.atclientlocation or _default('atclientlocation'))
            arrangement.atotherlocation = request.form.get(f'{day}_other',
                                                           arrangement.atotherlocation or _default('atotherlocation'))

        # Special conditions (stored on the contract record if present)
        raw = request.form.get('SpecialConditions')
        specialconditions = raw.strip() if isinstance(raw, str) else ''

        contract_record = db.session.scalar(
            select(ContractModel).where(ContractModel.sid == service_id)
        )
        if not contract_record:
            contract_record = ContractModel(sid=service_id, specialconditions=specialconditions)
            db.session.add(contract_record)
        else:
            contract_record.specialconditions = specialconditions

        # Commit all changes once
        db.session.commit()

        # Persist to session for later exports
        session['specialConditions'] = specialconditions
    
    # GET: (re)load for display
    arr_list = loadServiceArrangements(service_id) or []  # returns list and sets session cache      
    # Make a mapping keyed by day so Jinja can do arrangements.get("Monday")
    arrangements = {
        (row.get('day') if isinstance(row, dict) else row.day): row
        for row in arr_list
    }
    contract_record = db.session.scalar(
        select(ContractModel).where(ContractModel.sid == service_id)
    )
    if contract_record:
        contract['specialconditions'] = contract_record.specialconditions or ''
    return render_template('arrangements.html', arrangements=arrangements, contract=contract)


@views_bp.route('/clientcontract', methods=['GET', 'POST'])
def prepare_client_contract():

    contract = session.get('sessionContract', {})
    service_id = contract.get("sid", "")
    if service_id is None or service_id == "":
        flash("Select a Service Provider with a Service ID before continuing.", "error")
        return redirect(url_for('views.index'))

    # format start and end dates so they will populate date picker input correctly
    # first check if the date is in the expected format
    if contract and contract.get('startdate'):
        startdate_obj = datetime.strptime(contract['startdate'], "%d/%m/%Y")
        contract['startdate'] = startdate_obj.strftime("%Y-%m-%d")
    if contract and contract.get('enddate'):
        enddate_obj = datetime.strptime(contract['enddate'], "%d/%m/%Y")  
        contract['enddate'] = enddate_obj.strftime("%Y-%m-%d")
    return render_template('clientcontract.html', contract=contract)


@views_bp.route('/download_client_contract', methods=['POST'])
def download_client_contract():
    # Get session data
    contract = session.get('sessionContract', {})    
    sid = contract.get("sid", "")
    
    # Ensure service standards and arrangements are loaded
    service_standards = session.get('serviceStandards', loadServiceStandards(sid))    
    arrangements = session.get('serviceArrangements', loadServiceArrangements(sid))
    
    agreement_date = request.form.get('AgreementDate', '')
    f_agreement_date = datetime.strptime(agreement_date, "%Y-%m-%d").date()
    f_agreement_date = f_agreement_date.strftime("%d/%m/%Y")
        
    contract_record = db.session.scalar(select(ContractModel).where(ContractModel.sid == sid))
    special_conditions = contract_record.specialconditions if contract_record else ''
    
    # Build rows
    data_rows = []
    row = {}
    row['AgreementDate'] = f_agreement_date

    # Populate row with contract fields
    # making any neccessary substituions
    fields = ["companyname", "companyaddress", "companyjurisdiction", "companyregistrationnumber", 
              "sid", "servicename", "charges", 
              "contactname", "contacttitle", "contactemail", "contactphone", "contactaddress", 
              "startdate", "enddate", "duration", 
              "noticeperiod", "noticeperiod_unit",
              "dmname", "dmtitle", "dmemail", "dmphone"]

    export_columns = ["ClientName", "ClientAddress", "Jursidiction", "ClientCompanyNo",                       
                      "ServiceID", "ServiceName", "ClientCharge", 
                      "ContactName", "ContactTitle", "ContactEmail", "ContactPhone", "ContactAddress",  
                      "ServiceStart", "ServiceEnd", "Duration", 
                      "NoticePeriod", "NoticeUOM",
                      "dmname", "dmtitle", "dmemail", "dmphone"]
    
    for raw_field, column_name in zip(fields, export_columns):

        if raw_field == "companyjurisdiction":
            tmp_field = contract.get(raw_field,"")
            if tmp_field.strip().lower() == "england-wales":
                field = "England and Wales"
        else:
            field = contract.get(raw_field, '')
        field = contract.get(raw_field, '')
        # Tech debt: hard coded AD details        
        if (raw_field == "dmphone" ):
            field = "01379 871144"    
        row[column_name] = field

    row["SpecialConditions"] = special_conditions

    std_fields = ["ssn", "description"]
    std_export_columns = ["SSN", "SSDescription"]
    
    # Flatten service standards
    for i in range(10):
                
        # add the standard fields
        if i < len(service_standards):
            standard = service_standards[i]        
        else: 
            standard = {}

        for field, column_name in zip(std_fields, std_export_columns):
            f_column_name = f"{column_name}{i}"

            if standard:
                row[f_column_name] = standard.get(field, '')
            else:
                row[f_column_name] = ''

    # Flatten arrangements
    for arr in arrangements:
        day_string = arr['day'][:3]  # Get first 3 letters of the day
        for k, v in arr.items():
            if k == 'atclientlocation':
                row[f"ACL{day_string}"] = v
            elif k == 'atotherlocation':
                row[f"AOL{day_string}"] = v
            elif k == 'atservicebase':
                row[f"ASB{day_string}"] = v
            elif k == 'defaultserviceperiod':
                row[f"DSP{day_string}"] = v
                
    data_rows.append(row)

    # Create DataFrame
    df = pd.DataFrame(data_rows)

    # Write to Excel in-memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')

    output.seek(0)
    wb = load_workbook(output)
    ws = wb['Sheet1']

    # add table
    df_rows = len(df) + 1
    df_cols = len(df.columns)
    df_range = f"A1:{get_column_letter(df_cols)}{df_rows}"

    table1 = Table(displayName="Table1", ref=df_range)
    table1.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=False, showColumnStripes=False
    )
    ws.add_table(table1)

    # Save final output
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    
    # Upload to SharePoint
    target_url = "Uploads"
    download_name=f"{sid} Client Statement of Service CSOS.xlsx"

    file_bytes = final_output.getvalue()
    uploaded_file = uploadToSharePoint(file_bytes, download_name, target_url)

    if uploaded_file not in (200, 201):  
        flash(f"Failed to upload Client Statement of Service to SharePoint folder {target_url}. Error code {uploaded_file}", "error")
        return send_file(
            final_output,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    else:
        flash(f"Client Statement of Service uploaded to SharePoint.", "success")
        return redirect(url_for('views.index'))


@views_bp.route('/spmsa', methods=['GET', 'POST'])
def prepare_sp_msa():
    
    session_contract = session.get('sessionContract') or None

    if not session_contract:
        flash("Select a Service Provider before continuing.", "error")
        return redirect(url_for('views.index'))

    return render_template('spmsa.html', contract=session_contract)


@views_bp.route('/spnda', methods=['GET', 'POST'])
def prepare_sp_nda():

    session_contract = session.get('sessionContract') or None

    if not session_contract:
        flash("Select a Service Provider before continuing.", "error")
        return redirect(url_for('views.index'))

    return render_template('spnda.html', contract=session_contract)


@views_bp.route('/download_sp_msa', methods=['POST'])
def download_sp_msa():

    # Get session data
    contract = session.get('sessionContract', {})
    if not contract:
        flash("Select a Service Provider before continuing.", "error")
        return redirect(url_for('views.index'))

    agreement_date = request.form.get('AgreementDate', '')
    
    f_agreement_date = datetime.strptime(agreement_date, "%Y-%m-%d").date()
    f_agreement_date = f_agreement_date.strftime("%d/%m/%Y")

    # Build rows
    data_rows = []
    row = {}
    row["AgreementDate"] = f_agreement_date
    fields = ["candidateName", "candidateltdname", "candidatejurisdiction", "candidateltdregno", "candidateaddress", "candidateemail"]

    export_columns = ["CandidateName", "CandidateLtdName", "CandidateJurisdiction", "CandidateLtdRegNo", "CandidateAddress", "CandidateEmail"]

    # Populate row with contract fields
    # making any neccessary substitutions
    for raw_field, column_name in zip(fields, export_columns):
        if ( raw_field == "candidatejurisdiction" and contract.get(raw_field,'').strip().lower() == "england-wales" ):
            field = "England and Wales"
        elif ( raw_field == "candidateName" ):
            field = formatName(contract.get(raw_field,''))
        else:
            field = contract.get(raw_field, '')

        row[column_name] = field
            
    data_rows.append(row)

    # Create DataFrame
    df = pd.DataFrame(data_rows)

    # Write to Excel in-memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')

    output.seek(0)
    wb = load_workbook(output)
    ws = wb['Sheet1']

    # add table
    df_rows = len(df) + 1
    df_cols = len(df.columns)
    df_range = f"A1:{get_column_letter(df_cols)}{df_rows}"

    table1 = Table(displayName="Table1", ref=df_range)
    table1.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=False, showColumnStripes=False
    )
    ws.add_table(table1)

    # Save final output
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    # Upload to SharePoint
    target_url = "Uploads"
    download_name=f"{contract.get('candidateltdname')} Service Provider MSA SMSA.xlsx"

    file_bytes = final_output.getvalue()
    uploaded_file = uploadToSharePoint(file_bytes, download_name, target_url)

    if uploaded_file not in (200, 201):  
        flash(f"Failed to upload Service Provider MSA to SharePoint folder {target_url}. Error code {uploaded_file}", "error")
        return send_file(
            final_output,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    else:
        flash(f"Service Provider MSA uploaded to SharePoint.", "success")
        return redirect(url_for('views.index'))


@views_bp.route('/clientmsa', methods=['GET', 'POST'])
def prepare_client_msa():
    session_contract = session.get('sessionContract') or None

    if not session_contract:
        flash("Select a Service Provider before continuing.", "error")
        return redirect(url_for('views.index'))

    return render_template('clientmsa.html', contract=session_contract)


@views_bp.route('/download_client_msa', methods=['POST'])
def download_client_msa():
        
    # Get session data, obtain it from the form if necessary - most likely for an MSA
    contract = session.get('sessionContract', {})
    if not contract:
        client_name = request.form.get('clientname', '')
        client_id = request.form.get('clientId', '')    
        reg_no = request.form.get('RegNumber', '')
        jurisdiction = request.form.get('Jurisdiction', '')
        address = request.form.get('RegAddress', '')

        contract = {
            'companyname': client_name,
            'companyid': client_id,
            'companyregistrationnumber': reg_no,
            'companyjurisdiction': jurisdiction,
            'companyaddress': address,
            'contactname': request.form.get('contactname', ''),
            'contacttitle': request.form.get('contactTitle', '')
        }

    agreement_date = request.form.get('AgreementDate', '')

    f_agreement_date = ''

    if agreement_date != '':
        f_agreement_date = datetime.strptime(agreement_date, "%Y-%m-%d").date()
        f_agreement_date = f_agreement_date.strftime("%d/%m/%Y")

    # Build rows
    data_rows = []
    row = {}
    row["AgreementDate"] = f_agreement_date
    fields = ["companyname", "companyregistrationnumber", "companyjurisdiction", "companyaddress", "contactname", "contacttitle"]

    export_columns = ["ClientName", "CompanyNumber", "Jurisdiction", "CompanyAddress", "ContactName", "ContactTitle"]

    # Populate row with contract fields
    # making any neccessary substitutions
    for raw_field, column_name in zip(fields, export_columns):
        if ( raw_field == "companyjurisdiction" and contract.get(raw_field,'').strip().lower() == "england-wales" ):
            field = "England and Wales"
        elif ( raw_field == "candidateName" ):
            field = formatName(contract.get(raw_field,''))
        else:
            field = contract.get(raw_field, '')

        row[column_name] = field
            
    data_rows.append(row)

    # Create DataFrame
    df = pd.DataFrame(data_rows)

    # Write to Excel in-memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')

    output.seek(0)
    wb = load_workbook(output)
    ws = wb['Sheet1']

    # add table
    df_rows = len(df) + 1
    df_cols = len(df.columns)
    df_range = f"A1:{get_column_letter(df_cols)}{df_rows}"

    table1 = Table(displayName="Table1", ref=df_range)
    table1.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=False, showColumnStripes=False
    )
    ws.add_table(table1)

    # Save final output
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    
    # Upload to SharePoint

    target_url = "Uploads"
    download_name=f"{contract.get('companyname')} Client MSA CMSA.xlsx"

    file_bytes = final_output.getvalue()
    uploaded_file = uploadToSharePoint(file_bytes, download_name, target_url)

    if uploaded_file not in (200, 201):  
        flash(f"Failed to upload Client MSA to SharePoint folder {target_url}. Error code {uploaded_file}", "error")
        return send_file(
            final_output,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    else:
        flash(f"Client MSA uploaded to SharePoint folder {target_url}.", "success")
        return redirect(url_for('views.index'))
    

@views_bp.route('/chfetch', methods=['POST'])
def get_company_number():
    ltd_name = request.form.get('clientname','')
    data = {}
    if ltd_name:
        ch_result = searchCH(ltd_name)
        items = ch_result.get("items", [])
        match = next((item for item in items if item.get('title') == ltd_name.upper()), None)
        if match:
            data = {
                "regnumber": match.get("company_number", ""),
                "address": match.get("address_snippet", "")
            }
    return jsonify(data)
    

# --- Tiny in-memory cache to reduce API calls while typing ---
CACHE_TTL = 60  # seconds
_cache: dict[str, tuple[float, List[dict]]] = {}


def _cache_get(q: str) -> List[dict] | None:
    now = time.time()
    if q in _cache:
        ts, data = _cache[q]
        if now - ts <= CACHE_TTL:
            return data
        else:
            _cache.pop(q, None)
    return None


def _cache_set(q: str, data: List[dict]) -> None:
    _cache[q] = (time.time(), data)


def fetch_candidates(query: str) -> List[dict]:
    """
    Call the Colleague7 API to fetch candidates matching the query.
    """
    if not query:
        return []

    # Check cache first
    qkey = query.lower().strip()
    cached = _cache_get(qkey)
    if cached is not None:
        return cached

    payload = getC7Candidates(qkey)
    if payload is None or not isinstance(payload, list) or len(payload) == 0:
        return []   

    # Normalise to list[str] — tweak this for your API’s schema
    results = []
    
    for candidate_id in payload:

        candidate_data = getC7Candidate(candidate_id, query)
        if not candidate_data:
            continue

        candidate_dict = {
            "candidateId": candidate_data.get("candidateId",0) or candidate_data.get("candidateid",0),
            "candidateName": candidate_data.get("name",""),
            "candidateEmail": candidate_data.get("email",""),
            "candidatePhone": candidate_data.get("phone",""),
            "candidateLtdName": candidate_data.get("ltd_name",""),
            "candidateLtdRegNo": candidate_data.get("registration_number",""),
            "candidateJurisdiction": candidate_data.get("jurisdiction",""),
            "candidateAddress": candidate_data.get("address",""),
        }

        results.append(candidate_dict)

    # Update cache
    _cache_set(qkey, results)
    return results


def fetch_clients(query: str) -> List[dict]:
    """
    Normalise clients into a list of strings.
    C7 Company name search doesn't support partial matches, so we search on the Company class, first populating it if required
    """
    if not query:
        return []

    # Check cache first
    qkey = query.lower().strip()
    cached = _cache_get(qkey)
    if cached is not None:
        return cached

    payload = []
    if Company.count() == 0:
        payload = loadC7Clients()
    else:
        payload = Company.get_all_companies()
    
    results: List[dict] = []

    if isinstance(payload,list):
        if len(payload) == 0:
            return[]
        
        for client in payload:
            if client.companyname.lower().startswith(qkey):
                client_dict = {
                    "clientId": client.companyId,
                    "clientName": client.companyname,
                    "clientAddress": client.address,
                    "clientEmail": client.emailaddress,
                    "clientPhone": client.phone,
                    "clientRegNo": client.companyNumber,
                    "clientJurisdiction": client.jurisdiction                    
                }
                results.append(client_dict)

    # Update cache
    _cache_set(qkey, results)
    return results


def fetch_contacts(qclient: str, qcontact: str) -> List[dict]:
    """
    Normalise client contacts into a list of strings.
    """
    if not qclient:
        return []

    # Check cache first
    qkey = qcontact.lower().strip()
    cached = _cache_get(qkey)
    if cached is not None:
        return cached

    payload = getContactsByCompany(qclient)
    results: List[dict] = []

    if isinstance(payload, list):
        if len(payload) == 0:
            return []

        for clientcontact in payload:
            if clientcontact.get("ContactName", "").lower().startswith(qkey):
                contact_dict = {
                    "contactId": clientcontact.get("ContactId", ""),
                    "contactName": clientcontact.get("ContactName", ""),
                    "contactEmail": clientcontact.get("ContactEmail", ""),
                    "contactPhone": clientcontact.get("ContactPhone", ""),
                    "contactTitle": clientcontact.get("ContactTitle", ""),
                }
                results.append(contact_dict)

    # Update cache
    _cache_set(qkey, results)
    return results


def parse_date(value: str):
    try:
        # Parse the date string
        
        if len(value) >= 9:
            dt = datetime.strptime(value, "%a, %d %b %Y %H:%M:%S %Z")
        else:
            dt = datetime.strptime(value, "%a, %d %b")
        # return as a date object        
        date_only = dt.date()
        return date_only
    
    except (ValueError, TypeError):
        return None


def list_to_dict(prefix_list):
    result = {}
    for item in prefix_list:
        if ": " in item:
            key, value = item.split(": ", 1)
            result[key] = value
    return result


@views_bp.route('/candidatefetch', methods=['POST'])
def candidatefetch():
    candidate_name = request.form.get('CandidateName', '')
    result = searchC7Candidate(candidate_name) if candidate_name else None

    if not result:
        return jsonify({'error': 'No candidates found'}), 404
        
    return jsonify(result)


@views_bp.post("/contract/candidate")
def set_contract_candidate():
    data = request.get_json(force=True) or {}
    cand_id = data.get("candidateId")
    if not cand_id:
        return jsonify(error="candidateId required"), 400

    # contract = session.get("sessionContract", {})
        
    # Load existing contract data if available
    contract = gather_data(data)
    if contract:
        session['sessionContract'] = contract
        # only override session candidateName once, user must click 'Clear Session' to reset
        # this is to preserve the original candidate name including the service ID
        candidate_name = session.get("candidateName", "")
        if candidate_name is None or candidate_name == "":
            session["candidateName"] = contract.get("candidateName", "")    
        session.modified = True
    return ("", 204)


@views_bp.route('/download_sp_nda', methods=['POST'])
def download_sp_nda():

    # Get session data
    contract = session.get('sessionContract', {})
    if not contract:
        flash("Select a Service Provider before continuing.", "error")
        return redirect(url_for('views.index'))

    agreement_date = request.form.get('AgreementDate', '')
    
    f_agreement_date = datetime.strptime(agreement_date, "%Y-%m-%d").date()
    f_agreement_date = f_agreement_date.strftime("%d/%m/%Y")

    # Build rows
    data_rows = []
    row = {}
    row["AgreementDate"] = f_agreement_date
    fields = ["candidateName", "candidateaddress", "candidateemail"]
    
    export_columns = ["CandidateName", "CandidateAddress", "CandidateEmail"]
    
    # Populate row with contract fields    
    # making any neccessary substitutions
    for raw_field, column_name in zip(fields, export_columns):
        if ( raw_field == "candidateName" ):
            field = formatName(contract.get(raw_field,''))
        else:
            field = contract.get(raw_field, '')

        row[column_name] = field
            
    data_rows.append(row)

    # Create DataFrame
    df = pd.DataFrame(data_rows)

    # Write to Excel in-memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')

    output.seek(0)
    wb = load_workbook(output)
    ws = wb['Sheet1']

    # add table
    df_rows = len(df) + 1
    df_cols = len(df.columns)
    df_range = f"A1:{get_column_letter(df_cols)}{df_rows}"

    table1 = Table(displayName="Table1", ref=df_range)
    table1.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=False, showColumnStripes=False
    )
    ws.add_table(table1)

    # Save final output
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    # Upload to SharePoint
    target_url = "Uploads"
    download_name=f"{contract.get('candidateltdname')} Service Provider NDA SNDA.xlsx"

    file_bytes = final_output.getvalue()
    uploaded_file = uploadToSharePoint(file_bytes, download_name, target_url)

    if uploaded_file not in (200, 201):  
        flash(f"Failed to upload Service Provider NDA to SharePoint folder {target_url}. Error code {uploaded_file}", "error")
        return send_file(
            final_output,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    else:
        flash(f"Service Provider NDA uploaded to SharePoint.", "success")
        return redirect(url_for('views.index'))


@views_bp.route('/spcontract', methods=['GET', 'POST'])
def prepare_sp_contract():
    
    contract = session.get('sessionContract', {})
    service_id = contract.get("sid", "")
    if service_id is None or service_id == "":
        flash("Select a Service Provider with a Service ID before continuing.", "error")
        return redirect(url_for('views.index'))

    # format start and end dates so they will populate date picker input correctly
    # first check if the date is in the expected format
    if contract and contract.get('startdate'):
        startdate_obj = datetime.strptime(contract['startdate'], "%d/%m/%Y")
        contract['startdate'] = startdate_obj.strftime("%Y-%m-%d")
    if contract and contract.get('enddate'):
        enddate_obj = datetime.strptime(contract['enddate'], "%d/%m/%Y")  
        contract['enddate'] = enddate_obj.strftime("%Y-%m-%d")
    
    return render_template('spcontract.html', contract=contract)


@views_bp.route('/download_sp_contract', methods=['POST'])
def download_sp_contract():
    """
    Export SP contract data to Excel file for merge into docx
    There are 20 pairs of standard fields defined, this gives us a number of spares, should the number of CS or SP standards increase
    """
    # Get session data
    contract = session.get('sessionContract', {})    
    
    service_id = contract.get("sid", "")

    cs_standards = loadServiceStandards("CS") or []

    # Ensure service standards and arrangements are loaded
    service_standards = session.get('serviceStandards', loadServiceStandards(service_id))
    arrangements = session.get('serviceArrangements', loadServiceArrangements(service_id))
        
    agreement_date = request.form.get('AgreementDate', '')
    
    f_agreement_date = datetime.strptime(agreement_date, "%Y-%m-%d").date()
    f_agreement_date = f_agreement_date.strftime("%d/%m/%Y")

    contract_record = db.session.scalar(select(ContractModel).where(ContractModel.sid == service_id))
    special_conditions = contract_record.specialconditions if contract_record else ''

    # Build rows
    data_rows = []
    row = {}
    row['AgreementDate'] = f_agreement_date

    # Populate row with contract fields
    # making any neccessary substituions
    fields = ["companyname", "companyaddress", "companyjurisdiction", "companyregistrationnumber", 
              "sid", "servicename", "fees", 
              "contactname", "contacttitle", "contactemail", "contactphone", "contactaddress", 
              "startdate", "enddate", "duration", "noticeperiod", "noticeperiod_unit",
              "dmname", "dmtitle", "dmemail", "dmphone"]

    export_columns = ["ClientName", "ClientAddress", "Jursidiction", "ClientCompanyNo",                       
                      "ServiceID", "ServiceName", "Fees", 
                      "ContactName", "ContactTitle", "ContactEmail", "ContactPhone", "ContactAddress",  
                      "ServiceStart", "ServiceEnd", "Duration", "NoticePeriod", "NoticeUOM",
                      "dmname", "dmtitle", "dmemail", "dmphone"]

    for raw_field, column_name in zip(fields, export_columns):

        field = ""
        if raw_field == "companyjurisdiction":
            tmp_field = contract.get(raw_field,"")
            if tmp_field.strip().lower() == "england-wales":
                field = "England and Wales"
        else:
            field = contract.get(raw_field, '')
        # Tech debt: hard coded AD details        
        if (raw_field == "dmphone" ):
            field = "01379 871144"    
        row[column_name] = field

    row["SpecialConditions"] = special_conditions

    std_fields = ["ssn", "description"]
    std_export_columns = ["SSN", "SSDescription"]
    cs_count = 0
    
    # Flatten CS standards   
    for i, std in enumerate(cs_standards, start=1):        
        row[f"SSN{i}"] = std.ssn or ""
        row[f"SSDescription{i}"] = std.description or ""
        cs_count += 1

    # Flatten service standards
    for i in range(cs_count,20):

        i_active = i - cs_count + 1
        # add the standard fields
        if i_active < len(service_standards):
            standard = service_standards[i_active]        
        else: 
            standard = {}

        for field, column_name in zip(std_fields, std_export_columns):
            f_column_name = f"{column_name}{i}"

            if standard:
                row[f_column_name] = standard.get(field, '')
            else:
                row[f_column_name] = ''

    # Flatten arrangements
    for arr in arrangements:
        day_string = arr['day'][:3]  # Get first 3 letters of the day
        for k, v in arr.items():
            if k == 'atclientlocation':
                row[f"ACL{day_string}"] = v
            elif k == 'atotherlocation':
                row[f"AOL{day_string}"] = v
            elif k == 'atservicebase':
                row[f"ASB{day_string}"] = v
            elif k == 'defaultserviceperiod':
                row[f"DSP{day_string}"] = v
                
    data_rows.append(row)

    # Create DataFrame
    df = pd.DataFrame(data_rows)

    # Write to Excel in-memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')

    output.seek(0)
    wb = load_workbook(output)
    ws = wb['Sheet1']

    # add table
    df_rows = len(df) + 1
    df_cols = len(df.columns)
    df_range = f"A1:{get_column_letter(df_cols)}{df_rows}"

    table1 = Table(displayName="Table1", ref=df_range)
    table1.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=False, showColumnStripes=False
    )
    ws.add_table(table1)

    # Save final output
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    
        # Upload to SharePoint
    target_url = "Uploads"
    download_name=f"{service_id} Service Provider Statement of Service SSOS.xlsx"

    file_bytes = final_output.getvalue()
    uploaded_file = uploadToSharePoint(file_bytes, download_name, target_url)

    if uploaded_file not in (200, 201):  
        flash(f"Failed to upload Service Provider Statement of Service to SharePoint folder {target_url}. Error code {uploaded_file}", "error")
        return send_file(
            final_output,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    else:
        flash(f"Service Provider Statement of Service uploaded to SharePoint.", "success")
        return redirect(url_for('views.index'))